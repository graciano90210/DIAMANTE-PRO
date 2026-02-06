"""
Blueprint de Reportes - Diamante Pro
Dashboard de Inteligencia de Negocios (Mejorado)
"""
from flask import Blueprint, render_template, request, redirect, url_for, session, jsonify, Response, make_response
from datetime import datetime, timedelta
from ..models import Cliente, Prestamo, Pago, Transaccion, Ruta, Usuario, AporteCapital, Oficina, db
from sqlalchemy import func, case, and_, or_, extract
from decimal import Decimal
import json
import io

# Importar openpyxl para Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

reportes_bp = Blueprint('reportes', __name__)


def login_required(f):
    """Decorador para requerir login"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function


def calcular_tendencias_optimizado(fecha_inicio_tendencia, fecha_fin_tendencia, rol, usuario_id, oficina_id, rutas_oficina_ids, ruta_id):
    """
    Calcula tendencias de forma optimizada con una sola query agrupada por fecha.
    Esto reemplaza las 30 queries individuales por 2 queries agrupadas.
    """
    # Query optimizada para cobros por día
    cobros_query = db.session.query(
        func.date(Pago.fecha_pago).label('fecha'),
        func.coalesce(func.sum(Pago.monto), 0).label('total')
    ).filter(
        func.date(Pago.fecha_pago) >= fecha_inicio_tendencia,
        func.date(Pago.fecha_pago) <= fecha_fin_tendencia
    )

    if rol == 'cobrador':
        cobros_query = cobros_query.join(Prestamo).filter(Prestamo.cobrador_id == usuario_id)
    elif oficina_id and rutas_oficina_ids:
        cobros_query = cobros_query.join(Prestamo).filter(Prestamo.ruta_id.in_(rutas_oficina_ids))
    elif ruta_id:
        cobros_query = cobros_query.join(Prestamo).filter(Prestamo.ruta_id == ruta_id)

    cobros_query = cobros_query.group_by(func.date(Pago.fecha_pago))
    cobros_por_fecha = {str(r.fecha): float(r.total) for r in cobros_query.all()}

    # Query optimizada para préstamos por día
    prestamos_query = db.session.query(
        func.date(Prestamo.fecha_inicio).label('fecha'),
        func.coalesce(func.sum(Prestamo.monto_prestado), 0).label('total')
    ).filter(
        func.date(Prestamo.fecha_inicio) >= fecha_inicio_tendencia,
        func.date(Prestamo.fecha_inicio) <= fecha_fin_tendencia
    )

    if rol == 'cobrador':
        prestamos_query = prestamos_query.filter(Prestamo.cobrador_id == usuario_id)
    elif oficina_id and rutas_oficina_ids:
        prestamos_query = prestamos_query.filter(Prestamo.ruta_id.in_(rutas_oficina_ids))
    elif ruta_id:
        prestamos_query = prestamos_query.filter(Prestamo.ruta_id == ruta_id)

    prestamos_query = prestamos_query.group_by(func.date(Prestamo.fecha_inicio))
    prestamos_por_fecha = {str(r.fecha): float(r.total) for r in prestamos_query.all()}

    return cobros_por_fecha, prestamos_por_fecha


def calcular_heatmap_semanal(rol, usuario_id, oficina_id, rutas_oficina_ids, ruta_id):
    """
    Calcula el heatmap de cobros por día de la semana.
    Retorna datos para visualizar efectividad de cobro por día.
    """
    # Query para cobros agrupados por día de la semana (últimos 90 días)
    fecha_inicio = datetime.now().date() - timedelta(days=90)

    # SQLite usa strftime, PostgreSQL usa extract
    try:
        # Intentar con extract (PostgreSQL)
        cobros_query = db.session.query(
            extract('dow', Pago.fecha_pago).label('dia_semana'),
            func.count(Pago.id).label('cantidad'),
            func.coalesce(func.sum(Pago.monto), 0).label('total')
        ).filter(
            func.date(Pago.fecha_pago) >= fecha_inicio
        )
    except Exception:
        # Fallback para SQLite
        cobros_query = db.session.query(
            func.strftime('%w', Pago.fecha_pago).label('dia_semana'),
            func.count(Pago.id).label('cantidad'),
            func.coalesce(func.sum(Pago.monto), 0).label('total')
        ).filter(
            func.date(Pago.fecha_pago) >= fecha_inicio
        )

    if rol == 'cobrador':
        cobros_query = cobros_query.join(Prestamo).filter(Prestamo.cobrador_id == usuario_id)
    elif oficina_id and rutas_oficina_ids:
        cobros_query = cobros_query.join(Prestamo).filter(Prestamo.ruta_id.in_(rutas_oficina_ids))
    elif ruta_id:
        cobros_query = cobros_query.join(Prestamo).filter(Prestamo.ruta_id == ruta_id)

    try:
        cobros_query = cobros_query.group_by(extract('dow', Pago.fecha_pago))
    except Exception:
        cobros_query = cobros_query.group_by(func.strftime('%w', Pago.fecha_pago))

    dias = ['Domingo', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
    heatmap_data = {i: {'dia': dias[i], 'cantidad': 0, 'total': 0} for i in range(7)}

    for r in cobros_query.all():
        dia = int(r.dia_semana) if r.dia_semana is not None else 0
        heatmap_data[dia] = {
            'dia': dias[dia],
            'cantidad': r.cantidad,
            'total': float(r.total)
        }

    return list(heatmap_data.values())


def calcular_ciclo_vida_prestamos(filtro_prestamos):
    """
    Analiza el ciclo de vida de los préstamos:
    - Tiempo promedio de pago
    - Distribución por antigüedad
    - Tasa de renovación
    """
    hoy = datetime.now().date()

    # Préstamos por antigüedad (activos)
    prestamos = Prestamo.query.filter(*filtro_prestamos).all()

    antiguedad_0_30 = 0
    antiguedad_31_60 = 0
    antiguedad_61_90 = 0
    antiguedad_90_plus = 0

    total_dias_activos = 0
    prestamos_con_fecha = 0

    for p in prestamos:
        if p.fecha_inicio:
            dias = (hoy - p.fecha_inicio.date() if hasattr(p.fecha_inicio, 'date') else hoy - p.fecha_inicio).days
            total_dias_activos += dias
            prestamos_con_fecha += 1

            if dias <= 30:
                antiguedad_0_30 += 1
            elif dias <= 60:
                antiguedad_31_60 += 1
            elif dias <= 90:
                antiguedad_61_90 += 1
            else:
                antiguedad_90_plus += 1

    promedio_dias = total_dias_activos / prestamos_con_fecha if prestamos_con_fecha > 0 else 0

    # Préstamos cancelados (para calcular tiempo promedio de finalización)
    prestamos_finalizados = Prestamo.query.filter(
        Prestamo.estado == 'CANCELADO'
    ).order_by(Prestamo.id.desc()).limit(100).all()

    tiempos_finalizacion = []
    for p in prestamos_finalizados:
        if p.fecha_inicio and p.fecha_fin:
            inicio = p.fecha_inicio.date() if hasattr(p.fecha_inicio, 'date') else p.fecha_inicio
            fin = p.fecha_fin.date() if hasattr(p.fecha_fin, 'date') else p.fecha_fin
            dias = (fin - inicio).days
            if dias > 0:
                tiempos_finalizacion.append(dias)

    tiempo_promedio_finalizacion = sum(tiempos_finalizacion) / len(tiempos_finalizacion) if tiempos_finalizacion else 0

    return {
        'antiguedad_0_30': antiguedad_0_30,
        'antiguedad_31_60': antiguedad_31_60,
        'antiguedad_61_90': antiguedad_61_90,
        'antiguedad_90_plus': antiguedad_90_plus,
        'promedio_dias_activo': round(promedio_dias, 1),
        'tiempo_promedio_finalizacion': round(tiempo_promedio_finalizacion, 1)
    }


def calcular_proyeccion_cobros(prestamos_activos, dias_proyeccion=7):
    """
    Proyecta los cobros esperados para los próximos días.
    """
    hoy = datetime.now().date()
    proyeccion = []

    for i in range(dias_proyeccion):
        fecha = hoy + timedelta(days=i)
        dia_semana = fecha.weekday()
        dia_nombre = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom'][dia_semana]

        esperado = 0
        for p in prestamos_activos:
            cobra_hoy = False
            if p.frecuencia == 'DIARIO' and dia_semana != 6:
                cobra_hoy = True
            elif p.frecuencia == 'DIARIO_LUNES_VIERNES' and dia_semana < 5:
                cobra_hoy = True
            elif p.frecuencia == 'SEMANAL':
                # Asumir que cobra el mismo día de la semana que inició
                if p.fecha_inicio:
                    dia_inicio = p.fecha_inicio.weekday() if hasattr(p.fecha_inicio, 'weekday') else p.fecha_inicio.date().weekday()
                    if dia_semana == dia_inicio:
                        cobra_hoy = True
            elif p.frecuencia == 'BISEMANAL':
                cobra_hoy = dia_semana in [1, 4]  # Martes y Viernes

            if cobra_hoy:
                esperado += float(p.valor_cuota) if p.valor_cuota else 0

        proyeccion.append({
            'fecha': fecha.strftime('%d/%m'),
            'dia': dia_nombre,
            'esperado': esperado
        })

    return proyeccion


def calcular_metricas_periodo_anterior(fecha_inicio, fecha_fin, rol, usuario_id, oficina_id, rutas_oficina_ids, ruta_id):
    """
    Calcula métricas del período anterior para comparación.
    """
    duracion = (fecha_fin - fecha_inicio).days
    fecha_inicio_ant = fecha_inicio - timedelta(days=duracion + 1)
    fecha_fin_ant = fecha_inicio - timedelta(days=1)

    # Cobrado en período anterior
    pagos_ant_query = db.session.query(
        func.coalesce(func.sum(Pago.monto), 0)
    ).filter(
        Pago.fecha_pago >= fecha_inicio_ant,
        Pago.fecha_pago <= fecha_fin_ant
    )

    if rol == 'cobrador':
        pagos_ant_query = pagos_ant_query.join(Prestamo).filter(Prestamo.cobrador_id == usuario_id)
    elif oficina_id and rutas_oficina_ids:
        pagos_ant_query = pagos_ant_query.join(Prestamo).filter(Prestamo.ruta_id.in_(rutas_oficina_ids))
    elif ruta_id:
        pagos_ant_query = pagos_ant_query.join(Prestamo).filter(Prestamo.ruta_id == ruta_id)

    cobrado_anterior = float(pagos_ant_query.scalar() or 0)

    # Préstamos desembolsados período anterior
    prestamos_ant_query = db.session.query(
        func.coalesce(func.sum(Prestamo.monto_prestado), 0),
        func.count(Prestamo.id)
    ).filter(
        func.date(Prestamo.fecha_inicio) >= fecha_inicio_ant,
        func.date(Prestamo.fecha_inicio) <= fecha_fin_ant
    )

    if rol == 'cobrador':
        prestamos_ant_query = prestamos_ant_query.filter(Prestamo.cobrador_id == usuario_id)
    elif oficina_id and rutas_oficina_ids:
        prestamos_ant_query = prestamos_ant_query.filter(Prestamo.ruta_id.in_(rutas_oficina_ids))
    elif ruta_id:
        prestamos_ant_query = prestamos_ant_query.filter(Prestamo.ruta_id == ruta_id)

    result = prestamos_ant_query.first()
    desembolsado_anterior = float(result[0] or 0)
    cantidad_prestamos_anterior = result[1] or 0

    # Número de pagos período anterior
    num_pagos_ant_query = db.session.query(
        func.count(Pago.id)
    ).filter(
        Pago.fecha_pago >= fecha_inicio_ant,
        Pago.fecha_pago <= fecha_fin_ant
    )

    if rol == 'cobrador':
        num_pagos_ant_query = num_pagos_ant_query.join(Prestamo).filter(Prestamo.cobrador_id == usuario_id)
    elif oficina_id and rutas_oficina_ids:
        num_pagos_ant_query = num_pagos_ant_query.join(Prestamo).filter(Prestamo.ruta_id.in_(rutas_oficina_ids))
    elif ruta_id:
        num_pagos_ant_query = num_pagos_ant_query.join(Prestamo).filter(Prestamo.ruta_id == ruta_id)

    num_pagos_anterior = num_pagos_ant_query.scalar() or 0

    return {
        'cobrado_anterior': cobrado_anterior,
        'desembolsado_anterior': desembolsado_anterior,
        'cantidad_prestamos_anterior': cantidad_prestamos_anterior,
        'num_pagos_anterior': num_pagos_anterior,
        'fecha_inicio_anterior': fecha_inicio_ant.strftime('%Y-%m-%d'),
        'fecha_fin_anterior': fecha_fin_ant.strftime('%Y-%m-%d')
    }


def calcular_metricas_bi(fecha_inicio, fecha_fin, usuario_id=None, rol='dueno', ruta_id=None, oficina_id=None, cobrador_id=None, monto_min=None, monto_max=None, estado_filtro=None):
    """
    Calcula todas las métricas de Business Intelligence para el dashboard.
    Retorna un diccionario con todas las métricas calculadas.
    Versión mejorada con filtros avanzados y métricas adicionales.
    """
    hoy = datetime.now().date()

    # ==================== FILTROS BASE ====================
    # Estado del préstamo (por defecto activo, pero puede filtrarse)
    if estado_filtro and estado_filtro != 'TODOS':
        filtro_prestamos = [Prestamo.estado == estado_filtro]
    else:
        filtro_prestamos = [Prestamo.estado == 'ACTIVO']

    filtro_prestamos_todos = []
    filtro_pagos = []

    # Obtener IDs de rutas si hay oficina seleccionada
    rutas_oficina_ids = []
    if oficina_id:
        rutas_oficina_ids = [r.id for r in Ruta.query.filter_by(oficina_id=oficina_id).all()]

    if rol == 'cobrador':
        filtro_prestamos.append(Prestamo.cobrador_id == usuario_id)
        filtro_prestamos_todos.append(Prestamo.cobrador_id == usuario_id)
    elif oficina_id and rutas_oficina_ids:
        filtro_prestamos.append(Prestamo.ruta_id.in_(rutas_oficina_ids))
        filtro_prestamos_todos.append(Prestamo.ruta_id.in_(rutas_oficina_ids))
    elif ruta_id:
        filtro_prestamos.append(Prestamo.ruta_id == ruta_id)
        filtro_prestamos_todos.append(Prestamo.ruta_id == ruta_id)

    # Filtro por cobrador específico (nuevo)
    if cobrador_id:
        filtro_prestamos.append(Prestamo.cobrador_id == cobrador_id)
        filtro_prestamos_todos.append(Prestamo.cobrador_id == cobrador_id)

    # Filtro por rango de monto (nuevo)
    if monto_min is not None:
        filtro_prestamos.append(Prestamo.monto_prestado >= monto_min)
        filtro_prestamos_todos.append(Prestamo.monto_prestado >= monto_min)
    if monto_max is not None:
        filtro_prestamos.append(Prestamo.monto_prestado <= monto_max)
        filtro_prestamos_todos.append(Prestamo.monto_prestado <= monto_max)

    # ==================== MÉTRICAS DE CAPITAL ====================
    # Capital total prestado (préstamos activos)
    capital_prestado = db.session.query(
        func.coalesce(func.sum(Prestamo.monto_prestado), 0)
    ).filter(*filtro_prestamos).scalar() or 0

    # Cartera total (saldo pendiente por cobrar)
    cartera_total = db.session.query(
        func.coalesce(func.sum(Prestamo.saldo_actual), 0)
    ).filter(*filtro_prestamos).scalar() or 0

    # Capital recuperado (total cobrado histórico)
    capital_recuperado_query = db.session.query(
        func.coalesce(func.sum(Pago.monto), 0)
    )
    if rol == 'cobrador':
        capital_recuperado_query = capital_recuperado_query.join(Prestamo).filter(
            Prestamo.cobrador_id == usuario_id
        )
    elif oficina_id and rutas_oficina_ids:
        capital_recuperado_query = capital_recuperado_query.join(Prestamo).filter(
            Prestamo.ruta_id.in_(rutas_oficina_ids)
        )
    elif ruta_id:
        capital_recuperado_query = capital_recuperado_query.join(Prestamo).filter(
            Prestamo.ruta_id == ruta_id
        )
    capital_recuperado = capital_recuperado_query.scalar() or 0

    # ==================== GANANCIA NETA (Intereses) ====================
    # Ganancia = Cartera Total - Capital Prestado (lo que se gana por intereses)
    ganancia_esperada = float(cartera_total) - float(capital_prestado)

    # Ganancia cobrada en el período
    pagos_periodo_query = db.session.query(
        func.coalesce(func.sum(Pago.monto), 0)
    ).filter(
        Pago.fecha_pago >= fecha_inicio,
        Pago.fecha_pago <= fecha_fin
    )
    if rol == 'cobrador':
        pagos_periodo_query = pagos_periodo_query.join(Prestamo).filter(
            Prestamo.cobrador_id == usuario_id
        )
    elif oficina_id and rutas_oficina_ids:
        pagos_periodo_query = pagos_periodo_query.join(Prestamo).filter(
            Prestamo.ruta_id.in_(rutas_oficina_ids)
        )
    elif ruta_id:
        pagos_periodo_query = pagos_periodo_query.join(Prestamo).filter(
            Prestamo.ruta_id == ruta_id
        )
    cobrado_periodo = pagos_periodo_query.scalar() or 0

    # ==================== TASA DE MOROSIDAD ====================
    prestamos_activos = Prestamo.query.filter(*filtro_prestamos).all()
    total_prestamos_activos = len(prestamos_activos)

    prestamos_al_dia = sum(1 for p in prestamos_activos if p.cuotas_atrasadas == 0)
    prestamos_atrasados = sum(1 for p in prestamos_activos if p.cuotas_atrasadas > 0)
    prestamos_mora_grave = sum(1 for p in prestamos_activos if p.cuotas_atrasadas > 3)

    tasa_morosidad = (prestamos_atrasados / total_prestamos_activos * 100) if total_prestamos_activos > 0 else 0
    tasa_mora_grave = (prestamos_mora_grave / total_prestamos_activos * 100) if total_prestamos_activos > 0 else 0

    # Monto en mora
    monto_en_mora = sum(float(p.saldo_actual) for p in prestamos_activos if p.cuotas_atrasadas > 0)

    # ==================== FLUJO DE CAJA HOY ====================
    # Ingresos de hoy (pagos recibidos)
    ingresos_hoy_query = db.session.query(
        func.coalesce(func.sum(Pago.monto), 0)
    ).filter(func.date(Pago.fecha_pago) == hoy)
    if rol == 'cobrador':
        ingresos_hoy_query = ingresos_hoy_query.join(Prestamo).filter(
            Prestamo.cobrador_id == usuario_id
        )
    elif oficina_id and rutas_oficina_ids:
        ingresos_hoy_query = ingresos_hoy_query.join(Prestamo).filter(
            Prestamo.ruta_id.in_(rutas_oficina_ids)
        )
    elif ruta_id:
        ingresos_hoy_query = ingresos_hoy_query.join(Prestamo).filter(
            Prestamo.ruta_id == ruta_id
        )
    ingresos_hoy = ingresos_hoy_query.scalar() or 0

    # Egresos de hoy (préstamos desembolsados)
    egresos_hoy_query = db.session.query(
        func.coalesce(func.sum(Prestamo.monto_prestado), 0)
    ).filter(func.date(Prestamo.fecha_inicio) == hoy)
    if rol == 'cobrador':
        egresos_hoy_query = egresos_hoy_query.filter(Prestamo.cobrador_id == usuario_id)
    elif oficina_id and rutas_oficina_ids:
        egresos_hoy_query = egresos_hoy_query.filter(Prestamo.ruta_id.in_(rutas_oficina_ids))
    elif ruta_id:
        egresos_hoy_query = egresos_hoy_query.filter(Prestamo.ruta_id == ruta_id)
    egresos_hoy = egresos_hoy_query.scalar() or 0

    # Gastos operativos de hoy
    gastos_hoy_query = db.session.query(
        func.coalesce(func.sum(Transaccion.monto), 0)
    ).filter(
        func.date(Transaccion.fecha) == hoy,
        Transaccion.naturaleza == 'EGRESO'
    )
    if rol == 'cobrador':
        gastos_hoy_query = gastos_hoy_query.filter(Transaccion.usuario_origen_id == usuario_id)
    gastos_hoy = gastos_hoy_query.scalar() or 0

    flujo_caja_hoy = float(ingresos_hoy) - float(egresos_hoy) - float(gastos_hoy)

    # Por cobrar hoy (estimado según frecuencia)
    dia_semana = hoy.weekday()
    por_cobrar_hoy = 0
    for p in prestamos_activos:
        cobra_hoy = False
        if p.frecuencia == 'DIARIO' and dia_semana != 6:
            cobra_hoy = True
        elif p.frecuencia == 'DIARIO_LUNES_VIERNES' and dia_semana < 5:
            cobra_hoy = True
        elif p.frecuencia == 'BISEMANAL':
            cobra_hoy = True
        if cobra_hoy:
            por_cobrar_hoy += float(p.valor_cuota)

    tasa_cobro_hoy = (float(ingresos_hoy) / por_cobrar_hoy * 100) if por_cobrar_hoy > 0 else 0

    # ==================== TENDENCIAS (últimos 30 días) - OPTIMIZADO ====================
    fecha_inicio_tendencia = hoy - timedelta(days=29)
    cobros_por_fecha, prestamos_por_fecha = calcular_tendencias_optimizado(
        fecha_inicio_tendencia, hoy, rol, usuario_id, oficina_id, rutas_oficina_ids, ruta_id
    )

    tendencia_cobros = []
    tendencia_prestamos = []
    labels_tendencia = []

    for i in range(30):
        fecha = hoy - timedelta(days=29-i)
        fecha_str = str(fecha)
        labels_tendencia.append(fecha.strftime('%d/%m'))
        tendencia_cobros.append(cobros_por_fecha.get(fecha_str, 0))
        tendencia_prestamos.append(prestamos_por_fecha.get(fecha_str, 0))

    # ==================== HEATMAP SEMANAL (NUEVO) ====================
    heatmap_semanal = calcular_heatmap_semanal(rol, usuario_id, oficina_id, rutas_oficina_ids, ruta_id)

    # ==================== CICLO DE VIDA DE PRÉSTAMOS (NUEVO) ====================
    ciclo_vida = calcular_ciclo_vida_prestamos(filtro_prestamos)

    # ==================== PROYECCIÓN DE COBROS (NUEVO) ====================
    proyeccion_cobros = calcular_proyeccion_cobros(prestamos_activos, dias_proyeccion=7)

    # ==================== COMPARATIVA PERÍODO ANTERIOR (NUEVO) ====================
    periodo_anterior = calcular_metricas_periodo_anterior(
        fecha_inicio, fecha_fin, rol, usuario_id, oficina_id, rutas_oficina_ids, ruta_id
    )

    # ==================== DISTRIBUCIÓN DE CARTERA ====================
    # Por estado de morosidad
    cartera_al_dia = sum(float(p.saldo_actual) for p in prestamos_activos if p.cuotas_atrasadas == 0)
    cartera_atraso_leve = sum(float(p.saldo_actual) for p in prestamos_activos if 0 < p.cuotas_atrasadas <= 3)
    cartera_mora_grave = sum(float(p.saldo_actual) for p in prestamos_activos if p.cuotas_atrasadas > 3)

    # Por frecuencia
    distribucion_frecuencia_query = db.session.query(
        Prestamo.frecuencia,
        func.count(Prestamo.id).label('cantidad'),
        func.sum(Prestamo.saldo_actual).label('monto')
    ).filter(*filtro_prestamos).group_by(Prestamo.frecuencia).all()

    distribucion_frecuencia = [
        {'frecuencia': f[0], 'cantidad': f[1], 'monto': float(f[2] or 0)}
        for f in distribucion_frecuencia_query
    ]

    # ==================== TOP MÉTRICAS ====================
    # Top 5 deudores (incluye cliente_id para drill-down)
    top_deudores = db.session.query(
        Cliente.id,
        Cliente.nombre,
        Prestamo.saldo_actual,
        Prestamo.cuotas_atrasadas
    ).join(Prestamo).filter(*filtro_prestamos).order_by(
        Prestamo.saldo_actual.desc()
    ).limit(5).all()

    # Cobros por cobrador (solo para admin)
    cobros_por_cobrador = []
    if rol in ['dueno', 'gerente']:
        cobros_por_cobrador = db.session.query(
            Usuario.nombre,
            func.count(Pago.id).label('num_pagos'),
            func.coalesce(func.sum(Pago.monto), 0).label('total_cobrado')
        ).join(Pago, Usuario.id == Pago.cobrador_id).filter(
            Pago.fecha_pago >= fecha_inicio,
            Pago.fecha_pago <= fecha_fin
        ).group_by(Usuario.nombre).order_by(func.sum(Pago.monto).desc()).all()

    # ==================== MÉTRICAS POR MONEDA ====================
    metricas_por_moneda = {}
    monedas = db.session.query(Prestamo.moneda).filter(*filtro_prestamos).distinct().all()
    for (moneda,) in monedas:
        moneda = moneda or 'COP'
        filtro_moneda = filtro_prestamos + [Prestamo.moneda == moneda]

        capital_m = db.session.query(func.coalesce(func.sum(Prestamo.monto_prestado), 0)).filter(*filtro_moneda).scalar() or 0
        cartera_m = db.session.query(func.coalesce(func.sum(Prestamo.saldo_actual), 0)).filter(*filtro_moneda).scalar() or 0

        # Cartera por estado de morosidad por moneda
        prestamos_moneda = [p for p in prestamos_activos if (p.moneda or 'COP') == moneda]
        cartera_al_dia_m = sum(float(p.saldo_actual) for p in prestamos_moneda if p.cuotas_atrasadas == 0)
        cartera_atraso_leve_m = sum(float(p.saldo_actual) for p in prestamos_moneda if 0 < p.cuotas_atrasadas <= 3)
        cartera_mora_grave_m = sum(float(p.saldo_actual) for p in prestamos_moneda if p.cuotas_atrasadas > 3)
        monto_en_mora_m = sum(float(p.saldo_actual) for p in prestamos_moneda if p.cuotas_atrasadas > 0)

        # Ingresos de hoy por moneda
        ingresos_hoy_m_query = db.session.query(
            func.coalesce(func.sum(Pago.monto), 0)
        ).join(Prestamo).filter(
            func.date(Pago.fecha_pago) == hoy,
            Prestamo.moneda == moneda
        )
        if rol == 'cobrador':
            ingresos_hoy_m_query = ingresos_hoy_m_query.filter(Prestamo.cobrador_id == usuario_id)
        elif oficina_id and rutas_oficina_ids:
            ingresos_hoy_m_query = ingresos_hoy_m_query.filter(Prestamo.ruta_id.in_(rutas_oficina_ids))
        elif ruta_id:
            ingresos_hoy_m_query = ingresos_hoy_m_query.filter(Prestamo.ruta_id == ruta_id)
        ingresos_hoy_m = ingresos_hoy_m_query.scalar() or 0

        # Egresos de hoy por moneda
        egresos_hoy_m_query = db.session.query(
            func.coalesce(func.sum(Prestamo.monto_prestado), 0)
        ).filter(
            func.date(Prestamo.fecha_inicio) == hoy,
            Prestamo.moneda == moneda
        )
        if rol == 'cobrador':
            egresos_hoy_m_query = egresos_hoy_m_query.filter(Prestamo.cobrador_id == usuario_id)
        elif oficina_id and rutas_oficina_ids:
            egresos_hoy_m_query = egresos_hoy_m_query.filter(Prestamo.ruta_id.in_(rutas_oficina_ids))
        elif ruta_id:
            egresos_hoy_m_query = egresos_hoy_m_query.filter(Prestamo.ruta_id == ruta_id)
        egresos_hoy_m = egresos_hoy_m_query.scalar() or 0

        metricas_por_moneda[moneda] = {
            'capital': float(capital_m),
            'cartera': float(cartera_m),
            'ganancia': float(cartera_m) - float(capital_m),
            'cartera_al_dia': cartera_al_dia_m,
            'cartera_atraso_leve': cartera_atraso_leve_m,
            'cartera_mora_grave': cartera_mora_grave_m,
            'monto_en_mora': monto_en_mora_m,
            'ingresos_hoy': float(ingresos_hoy_m),
            'egresos_hoy': float(egresos_hoy_m),
            'flujo_neto_hoy': float(ingresos_hoy_m) - float(egresos_hoy_m)
        }

    # ==================== CALCULAR VARIACIONES % ====================
    variacion_cobros = 0
    variacion_prestamos = 0
    if periodo_anterior['cobrado_anterior'] > 0:
        variacion_cobros = round(((float(cobrado_periodo) - periodo_anterior['cobrado_anterior']) / periodo_anterior['cobrado_anterior']) * 100, 1)

    # Contar préstamos del período actual
    prestamos_periodo_query = db.session.query(func.count(Prestamo.id)).filter(
        func.date(Prestamo.fecha_inicio) >= fecha_inicio,
        func.date(Prestamo.fecha_inicio) <= fecha_fin
    )
    if rol == 'cobrador':
        prestamos_periodo_query = prestamos_periodo_query.filter(Prestamo.cobrador_id == usuario_id)
    elif oficina_id and rutas_oficina_ids:
        prestamos_periodo_query = prestamos_periodo_query.filter(Prestamo.ruta_id.in_(rutas_oficina_ids))
    elif ruta_id:
        prestamos_periodo_query = prestamos_periodo_query.filter(Prestamo.ruta_id == ruta_id)
    cantidad_prestamos_actual = prestamos_periodo_query.scalar() or 0

    if periodo_anterior['cantidad_prestamos_anterior'] > 0:
        variacion_prestamos = round(((cantidad_prestamos_actual - periodo_anterior['cantidad_prestamos_anterior']) / periodo_anterior['cantidad_prestamos_anterior']) * 100, 1)

    # ==================== ALERTAS Y METAS ====================
    alertas = []

    # Alerta de morosidad alta
    if tasa_morosidad > 30:
        alertas.append({
            'tipo': 'danger',
            'icono': 'exclamation-triangle-fill',
            'titulo': 'Morosidad Crítica',
            'mensaje': f'La tasa de morosidad ({tasa_morosidad}%) supera el 30%'
        })
    elif tasa_morosidad > 20:
        alertas.append({
            'tipo': 'warning',
            'icono': 'exclamation-circle-fill',
            'titulo': 'Morosidad Alta',
            'mensaje': f'La tasa de morosidad ({tasa_morosidad}%) supera el 20%'
        })

    # Alerta de tasa de cobro baja
    if tasa_cobro_hoy < 50 and por_cobrar_hoy > 0:
        alertas.append({
            'tipo': 'warning',
            'icono': 'cash-stack',
            'titulo': 'Cobro Bajo Hoy',
            'mensaje': f'Solo se ha cobrado el {tasa_cobro_hoy}% de lo esperado'
        })

    # Alerta de mora grave
    if prestamos_mora_grave > 5:
        alertas.append({
            'tipo': 'danger',
            'icono': 'person-x-fill',
            'titulo': 'Mora Grave',
            'mensaje': f'{prestamos_mora_grave} préstamos con más de 3 cuotas atrasadas'
        })

    # Alerta positiva de buen desempeño
    if tasa_cobro_hoy >= 90 and por_cobrar_hoy > 0:
        alertas.append({
            'tipo': 'success',
            'icono': 'trophy-fill',
            'titulo': 'Excelente Cobro',
            'mensaje': f'¡{tasa_cobro_hoy}% de cobro alcanzado hoy!'
        })

    # Metas sugeridas basadas en histórico
    promedio_cobro_diario = sum(tendencia_cobros) / 30 if tendencia_cobros else 0
    meta_diaria_sugerida = promedio_cobro_diario * 1.1  # 10% más que el promedio

    return {
        # Capital
        'capital_prestado': float(capital_prestado),
        'cartera_total': float(cartera_total),
        'capital_recuperado': float(capital_recuperado),
        'ganancia_esperada': ganancia_esperada,
        'cobrado_periodo': float(cobrado_periodo),

        # Morosidad
        'total_prestamos_activos': total_prestamos_activos,
        'prestamos_al_dia': prestamos_al_dia,
        'prestamos_atrasados': prestamos_atrasados,
        'prestamos_mora_grave': prestamos_mora_grave,
        'tasa_morosidad': round(tasa_morosidad, 1),
        'tasa_mora_grave': round(tasa_mora_grave, 1),
        'monto_en_mora': monto_en_mora,

        # Flujo de caja
        'ingresos_hoy': float(ingresos_hoy),
        'egresos_hoy': float(egresos_hoy),
        'gastos_hoy': float(gastos_hoy),
        'flujo_caja_hoy': flujo_caja_hoy,
        'por_cobrar_hoy': por_cobrar_hoy,
        'tasa_cobro_hoy': round(tasa_cobro_hoy, 1),

        # Tendencias
        'labels_tendencia': labels_tendencia,
        'tendencia_cobros': tendencia_cobros,
        'tendencia_prestamos': tendencia_prestamos,

        # Distribución
        'cartera_al_dia': cartera_al_dia,
        'cartera_atraso_leve': cartera_atraso_leve,
        'cartera_mora_grave': cartera_mora_grave,
        'distribucion_frecuencia': distribucion_frecuencia,

        # Top
        'top_deudores': [{'id': d[0], 'nombre': d[1], 'saldo': float(d[2]), 'atraso': d[3]} for d in top_deudores],
        'cobros_por_cobrador': [{'nombre': c[0], 'pagos': c[1], 'total': float(c[2])} for c in cobros_por_cobrador],

        # Por moneda
        'metricas_por_moneda': metricas_por_moneda,

        # ==================== NUEVAS MÉTRICAS ====================
        # Heatmap semanal
        'heatmap_semanal': heatmap_semanal,

        # Ciclo de vida
        'ciclo_vida': ciclo_vida,

        # Proyección de cobros
        'proyeccion_cobros': proyeccion_cobros,

        # Comparativa período anterior
        'periodo_anterior': periodo_anterior,
        'variacion_cobros': variacion_cobros,
        'variacion_prestamos': variacion_prestamos,
        'cantidad_prestamos_actual': cantidad_prestamos_actual,

        # Alertas
        'alertas': alertas,

        # Metas
        'promedio_cobro_diario': round(promedio_cobro_diario, 0),
        'meta_diaria_sugerida': round(meta_diaria_sugerida, 0)
    }


@reportes_bp.route('/reportes')
@login_required
def reportes():
    """Dashboard de Inteligencia de Negocios - Versión Mejorada"""
    usuario_id = session.get('usuario_id')
    rol = session.get('rol')

    # Obtener filtros de ruta y oficina (priorizar parámetros GET sobre sesión)
    ruta_seleccionada_id = request.args.get('ruta_id', type=int)
    oficina_seleccionada_id = request.args.get('oficina_id', type=int)

    # Nuevos filtros avanzados
    cobrador_seleccionado_id = request.args.get('cobrador_id', type=int)
    monto_min = request.args.get('monto_min', type=float)
    monto_max = request.args.get('monto_max', type=float)
    estado_filtro = request.args.get('estado', default='ACTIVO')

    # Si no hay parámetros GET, usar los de la sesión como fallback
    if ruta_seleccionada_id is None and oficina_seleccionada_id is None:
        ruta_seleccionada_id = session.get('ruta_seleccionada_id')
        oficina_seleccionada_id = session.get('oficina_seleccionada_id')

    # Obtener fecha de inicio y fin (por defecto últimos 30 días)
    fecha_fin = datetime.now()
    fecha_inicio = fecha_fin - timedelta(days=30)

    if request.args.get('fecha_inicio'):
        fecha_inicio = datetime.strptime(request.args.get('fecha_inicio'), '%Y-%m-%d')
    if request.args.get('fecha_fin'):
        fecha_fin = datetime.strptime(request.args.get('fecha_fin'), '%Y-%m-%d')
        fecha_fin = fecha_fin.replace(hour=23, minute=59, second=59)

    # Calcular métricas de BI con filtros avanzados
    metricas = calcular_metricas_bi(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        usuario_id=usuario_id,
        rol=rol,
        ruta_id=ruta_seleccionada_id,
        oficina_id=oficina_seleccionada_id,
        cobrador_id=cobrador_seleccionado_id,
        monto_min=monto_min,
        monto_max=monto_max,
        estado_filtro=estado_filtro
    )

    # Estadísticas generales adicionales
    total_clientes = Cliente.query.count()
    total_prestamos = Prestamo.query.count()
    prestamos_cancelados = Prestamo.query.filter_by(estado='CANCELADO').count()

    # Cargar rutas y oficinas para filtros
    todas_las_rutas = Ruta.query.order_by(Ruta.nombre).all()
    todas_las_oficinas = Oficina.query.filter_by(activo=True).order_by(Oficina.nombre).all()

    # Cargar cobradores para filtro
    todos_los_cobradores = Usuario.query.filter(
        Usuario.rol.in_(['cobrador', 'gerente', 'dueno'])
    ).order_by(Usuario.nombre).all()

    ruta_seleccionada = None
    if ruta_seleccionada_id:
        ruta_seleccionada = Ruta.query.get(ruta_seleccionada_id)

    oficina_seleccionada = None
    if oficina_seleccionada_id:
        oficina_seleccionada = Oficina.query.get(oficina_seleccionada_id)

    cobrador_seleccionado = None
    if cobrador_seleccionado_id:
        cobrador_seleccionado = Usuario.query.get(cobrador_seleccionado_id)

    return render_template('reportes_bi.html',
        fecha_inicio=fecha_inicio.strftime('%Y-%m-%d'),
        fecha_fin=fecha_fin.strftime('%Y-%m-%d'),
        metricas=metricas,
        total_clientes=total_clientes,
        total_prestamos=total_prestamos,
        prestamos_cancelados=prestamos_cancelados,
        todas_las_rutas=todas_las_rutas,
        todas_las_oficinas=todas_las_oficinas,
        todos_los_cobradores=todos_los_cobradores,
        ruta_seleccionada=ruta_seleccionada,
        oficina_seleccionada=oficina_seleccionada,
        cobrador_seleccionado=cobrador_seleccionado,
        monto_min=monto_min,
        monto_max=monto_max,
        estado_filtro=estado_filtro,
        nombre=session.get('nombre'),
        rol=session.get('rol'))


@reportes_bp.route('/api/reportes/metricas')
@login_required
def api_metricas():
    """API endpoint para obtener métricas en tiempo real (AJAX)"""
    usuario_id = session.get('usuario_id')
    rol = session.get('rol')
    ruta_id = session.get('ruta_seleccionada_id')
    oficina_id = session.get('oficina_seleccionada_id')

    fecha_fin = datetime.now()
    fecha_inicio = fecha_fin - timedelta(days=30)

    metricas = calcular_metricas_bi(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        usuario_id=usuario_id,
        rol=rol,
        ruta_id=ruta_id,
        oficina_id=oficina_id
    )

    return jsonify(metricas)


# ==================== EXPORTACIÓN A EXCEL ====================
@reportes_bp.route('/reportes/exportar/excel')
@login_required
def exportar_excel():
    """Exporta el reporte completo a Excel"""
    if not EXCEL_AVAILABLE:
        return jsonify({'error': 'Módulo openpyxl no disponible'}), 500

    usuario_id = session.get('usuario_id')
    rol = session.get('rol')

    # Obtener filtros
    ruta_id = request.args.get('ruta_id', type=int)
    oficina_id = request.args.get('oficina_id', type=int)
    cobrador_id = request.args.get('cobrador_id', type=int)

    fecha_fin = datetime.now()
    fecha_inicio = fecha_fin - timedelta(days=30)

    if request.args.get('fecha_inicio'):
        fecha_inicio = datetime.strptime(request.args.get('fecha_inicio'), '%Y-%m-%d')
    if request.args.get('fecha_fin'):
        fecha_fin = datetime.strptime(request.args.get('fecha_fin'), '%Y-%m-%d')

    # Calcular métricas
    metricas = calcular_metricas_bi(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        usuario_id=usuario_id,
        rol=rol,
        ruta_id=ruta_id,
        oficina_id=oficina_id,
        cobrador_id=cobrador_id
    )

    # Crear workbook
    wb = Workbook()

    # ===== Hoja 1: Resumen Ejecutivo =====
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0A0E21", end_color="0A0E21", fill_type="solid")
    cyan_fill = PatternFill(start_color="00E5FF", end_color="00E5FF", fill_type="solid")
    gold_font = Font(bold=True, color="FFD700", size=14)
    number_format = '#,##0'

    # Título
    ws1['A1'] = "REPORTE DE INTELIGENCIA DE NEGOCIOS - DIAMANTE PRO"
    ws1['A1'].font = Font(bold=True, size=16, color="00E5FF")
    ws1.merge_cells('A1:D1')

    ws1['A2'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
    ws1['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    # KPIs principales
    row = 5
    kpis = [
        ("Capital Prestado", metricas['capital_prestado']),
        ("Cartera Total", metricas['cartera_total']),
        ("Ganancia Esperada", metricas['ganancia_esperada']),
        ("Capital Recuperado", metricas['capital_recuperado']),
        ("Cobrado en Período", metricas['cobrado_periodo']),
        ("", ""),
        ("MOROSIDAD", ""),
        ("Tasa de Morosidad", f"{metricas['tasa_morosidad']}%"),
        ("Préstamos Activos", metricas['total_prestamos_activos']),
        ("Préstamos Al Día", metricas['prestamos_al_dia']),
        ("Préstamos Atrasados", metricas['prestamos_atrasados']),
        ("Préstamos en Mora Grave", metricas['prestamos_mora_grave']),
        ("Monto en Mora", metricas['monto_en_mora']),
        ("", ""),
        ("FLUJO DE CAJA HOY", ""),
        ("Ingresos Hoy", metricas['ingresos_hoy']),
        ("Egresos Hoy", metricas['egresos_hoy']),
        ("Gastos Hoy", metricas['gastos_hoy']),
        ("Flujo Neto", metricas['flujo_caja_hoy']),
        ("Tasa de Cobro Hoy", f"{metricas['tasa_cobro_hoy']}%"),
    ]

    ws1['A4'] = "MÉTRICAS CLAVE"
    ws1['A4'].font = header_font
    ws1['A4'].fill = header_fill
    ws1['B4'] = "VALOR"
    ws1['B4'].font = header_font
    ws1['B4'].fill = header_fill

    for i, (label, value) in enumerate(kpis, start=5):
        ws1[f'A{i}'] = label
        if isinstance(value, (int, float)) and label:
            ws1[f'B{i}'] = value
            ws1[f'B{i}'].number_format = number_format
        else:
            ws1[f'B{i}'] = value

    # Ajustar anchos
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20

    # ===== Hoja 2: Top Deudores =====
    ws2 = wb.create_sheet("Top Deudores")
    ws2['A1'] = "Nombre"
    ws2['B1'] = "Saldo"
    ws2['C1'] = "Cuotas Atrasadas"

    for col in ['A', 'B', 'C']:
        ws2[f'{col}1'].font = header_font
        ws2[f'{col}1'].fill = header_fill

    for i, deudor in enumerate(metricas['top_deudores'], start=2):
        ws2[f'A{i}'] = deudor['nombre']
        ws2[f'B{i}'] = deudor['saldo']
        ws2[f'B{i}'].number_format = number_format
        ws2[f'C{i}'] = deudor['atraso']

    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 18

    # ===== Hoja 3: Rendimiento Cobradores =====
    ws3 = wb.create_sheet("Rendimiento Cobradores")
    ws3['A1'] = "Cobrador"
    ws3['B1'] = "Pagos Realizados"
    ws3['C1'] = "Total Cobrado"
    ws3['D1'] = "Promedio por Pago"

    for col in ['A', 'B', 'C', 'D']:
        ws3[f'{col}1'].font = header_font
        ws3[f'{col}1'].fill = header_fill

    for i, cobrador in enumerate(metricas['cobros_por_cobrador'], start=2):
        ws3[f'A{i}'] = cobrador['nombre']
        ws3[f'B{i}'] = cobrador['pagos']
        ws3[f'C{i}'] = cobrador['total']
        ws3[f'C{i}'].number_format = number_format
        promedio = cobrador['total'] / cobrador['pagos'] if cobrador['pagos'] > 0 else 0
        ws3[f'D{i}'] = promedio
        ws3[f'D{i}'].number_format = number_format

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 18

    # ===== Hoja 4: Tendencia 30 días =====
    ws4 = wb.create_sheet("Tendencia 30 días")
    ws4['A1'] = "Fecha"
    ws4['B1'] = "Cobros"
    ws4['C1'] = "Préstamos"

    for col in ['A', 'B', 'C']:
        ws4[f'{col}1'].font = header_font
        ws4[f'{col}1'].fill = header_fill

    for i, (fecha, cobro, prestamo) in enumerate(zip(
        metricas['labels_tendencia'],
        metricas['tendencia_cobros'],
        metricas['tendencia_prestamos']
    ), start=2):
        ws4[f'A{i}'] = fecha
        ws4[f'B{i}'] = cobro
        ws4[f'B{i}'].number_format = number_format
        ws4[f'C{i}'] = prestamo
        ws4[f'C{i}'].number_format = number_format

    ws4.column_dimensions['A'].width = 12
    ws4.column_dimensions['B'].width = 18
    ws4.column_dimensions['C'].width = 18

    # ===== Hoja 5: Proyección de Cobros =====
    ws5 = wb.create_sheet("Proyección Cobros")
    ws5['A1'] = "Fecha"
    ws5['B1'] = "Día"
    ws5['C1'] = "Cobro Esperado"

    for col in ['A', 'B', 'C']:
        ws5[f'{col}1'].font = header_font
        ws5[f'{col}1'].fill = header_fill

    for i, proy in enumerate(metricas['proyeccion_cobros'], start=2):
        ws5[f'A{i}'] = proy['fecha']
        ws5[f'B{i}'] = proy['dia']
        ws5[f'C{i}'] = proy['esperado']
        ws5[f'C{i}'].number_format = number_format

    ws5.column_dimensions['A'].width = 12
    ws5.column_dimensions['B'].width = 12
    ws5.column_dimensions['C'].width = 18

    # ===== Hoja 6: Heatmap Semanal =====
    ws6 = wb.create_sheet("Efectividad por Día")
    ws6['A1'] = "Día"
    ws6['B1'] = "Cantidad Cobros"
    ws6['C1'] = "Total Cobrado"

    for col in ['A', 'B', 'C']:
        ws6[f'{col}1'].font = header_font
        ws6[f'{col}1'].fill = header_fill

    for i, dia in enumerate(metricas['heatmap_semanal'], start=2):
        ws6[f'A{i}'] = dia['dia']
        ws6[f'B{i}'] = dia['cantidad']
        ws6[f'C{i}'] = dia['total']
        ws6[f'C{i}'].number_format = number_format

    ws6.column_dimensions['A'].width = 15
    ws6.column_dimensions['B'].width = 18
    ws6.column_dimensions['C'].width = 18

    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_bi_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

    return response


# ==================== EXPORTACIÓN A CSV ====================
@reportes_bp.route('/reportes/exportar/csv')
@login_required
def exportar_csv():
    """Exporta el reporte a CSV"""
    usuario_id = session.get('usuario_id')
    rol = session.get('rol')

    # Obtener filtros
    ruta_id = request.args.get('ruta_id', type=int)
    oficina_id = request.args.get('oficina_id', type=int)

    fecha_fin = datetime.now()
    fecha_inicio = fecha_fin - timedelta(days=30)

    if request.args.get('fecha_inicio'):
        fecha_inicio = datetime.strptime(request.args.get('fecha_inicio'), '%Y-%m-%d')
    if request.args.get('fecha_fin'):
        fecha_fin = datetime.strptime(request.args.get('fecha_fin'), '%Y-%m-%d')

    # Calcular métricas
    metricas = calcular_metricas_bi(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        usuario_id=usuario_id,
        rol=rol,
        ruta_id=ruta_id,
        oficina_id=oficina_id
    )

    # Generar CSV
    output = io.StringIO()

    # Cabecera
    output.write("REPORTE DE INTELIGENCIA DE NEGOCIOS - DIAMANTE PRO\n")
    output.write(f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}\n")
    output.write(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n\n")

    # KPIs
    output.write("=== MÉTRICAS CLAVE ===\n")
    output.write(f"Capital Prestado,{metricas['capital_prestado']}\n")
    output.write(f"Cartera Total,{metricas['cartera_total']}\n")
    output.write(f"Ganancia Esperada,{metricas['ganancia_esperada']}\n")
    output.write(f"Capital Recuperado,{metricas['capital_recuperado']}\n")
    output.write(f"Cobrado en Período,{metricas['cobrado_periodo']}\n\n")

    output.write("=== MOROSIDAD ===\n")
    output.write(f"Tasa de Morosidad,{metricas['tasa_morosidad']}%\n")
    output.write(f"Préstamos Activos,{metricas['total_prestamos_activos']}\n")
    output.write(f"Préstamos Al Día,{metricas['prestamos_al_dia']}\n")
    output.write(f"Préstamos Atrasados,{metricas['prestamos_atrasados']}\n")
    output.write(f"Monto en Mora,{metricas['monto_en_mora']}\n\n")

    output.write("=== FLUJO DE CAJA HOY ===\n")
    output.write(f"Ingresos,{metricas['ingresos_hoy']}\n")
    output.write(f"Egresos,{metricas['egresos_hoy']}\n")
    output.write(f"Flujo Neto,{metricas['flujo_caja_hoy']}\n\n")

    # Top Deudores
    output.write("=== TOP DEUDORES ===\n")
    output.write("Nombre,Saldo,Cuotas Atrasadas\n")
    for d in metricas['top_deudores']:
        output.write(f"{d['nombre']},{d['saldo']},{d['atraso']}\n")
    output.write("\n")

    # Tendencia
    output.write("=== TENDENCIA 30 DÍAS ===\n")
    output.write("Fecha,Cobros,Préstamos\n")
    for fecha, cobro, prestamo in zip(
        metricas['labels_tendencia'],
        metricas['tendencia_cobros'],
        metricas['tendencia_prestamos']
    ):
        output.write(f"{fecha},{cobro},{prestamo}\n")

    # Proyección
    output.write("\n=== PROYECCIÓN PRÓXIMOS 7 DÍAS ===\n")
    output.write("Fecha,Día,Cobro Esperado\n")
    for p in metricas['proyeccion_cobros']:
        output.write(f"{p['fecha']},{p['dia']},{p['esperado']}\n")

    response = Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=reporte_bi_{datetime.now().strftime("%Y%m%d_%H%M")}.csv'}
    )

    return response


# ==================== API PARA DRILL-DOWN ====================
@reportes_bp.route('/api/reportes/deudores/<int:cliente_id>')
@login_required
def api_detalle_deudor(cliente_id):
    """Obtiene detalle de un deudor específico para drill-down"""
    cliente = Cliente.query.get_or_404(cliente_id)

    prestamos = Prestamo.query.filter_by(cliente_id=cliente_id).all()

    prestamos_data = []
    for p in prestamos:
        pagos = Pago.query.filter_by(prestamo_id=p.id).order_by(Pago.fecha_pago.desc()).limit(10).all()
        prestamos_data.append({
            'id': p.id,
            'monto_prestado': float(p.monto_prestado),
            'saldo_actual': float(p.saldo_actual),
            'cuotas_atrasadas': p.cuotas_atrasadas,
            'estado': p.estado,
            'fecha_inicio': p.fecha_inicio.strftime('%d/%m/%Y') if p.fecha_inicio else None,
            'ultimos_pagos': [{
                'fecha': pago.fecha_pago.strftime('%d/%m/%Y') if pago.fecha_pago else None,
                'monto': float(pago.monto)
            } for pago in pagos]
        })

    return jsonify({
        'cliente': {
            'id': cliente.id,
            'nombre': cliente.nombre,
            'telefono': cliente.telefono,
            'direccion': cliente.direccion
        },
        'prestamos': prestamos_data
    })


# ==================== API PARA GRÁFICOS EN TIEMPO REAL ====================
@reportes_bp.route('/api/reportes/grafico/<tipo>')
@login_required
def api_grafico(tipo):
    """Endpoints para actualizar gráficos específicos"""
    usuario_id = session.get('usuario_id')
    rol = session.get('rol')
    ruta_id = request.args.get('ruta_id', type=int)
    oficina_id = request.args.get('oficina_id', type=int)

    rutas_oficina_ids = []
    if oficina_id:
        rutas_oficina_ids = [r.id for r in Ruta.query.filter_by(oficina_id=oficina_id).all()]

    if tipo == 'heatmap':
        data = calcular_heatmap_semanal(rol, usuario_id, oficina_id, rutas_oficina_ids, ruta_id)
        return jsonify(data)

    elif tipo == 'proyeccion':
        filtro = [Prestamo.estado == 'ACTIVO']
        if rol == 'cobrador':
            filtro.append(Prestamo.cobrador_id == usuario_id)
        elif ruta_id:
            filtro.append(Prestamo.ruta_id == ruta_id)
        prestamos = Prestamo.query.filter(*filtro).all()
        data = calcular_proyeccion_cobros(prestamos)
        return jsonify(data)

    elif tipo == 'ciclo_vida':
        filtro = [Prestamo.estado == 'ACTIVO']
        if rol == 'cobrador':
            filtro.append(Prestamo.cobrador_id == usuario_id)
        elif ruta_id:
            filtro.append(Prestamo.ruta_id == ruta_id)
        data = calcular_ciclo_vida_prestamos(filtro)
        return jsonify(data)

    return jsonify({'error': 'Tipo de gráfico no válido'}), 400


# ==================== REPORTE POR RUTA / RECAUDADOR ====================
def calcular_metricas_ruta_diaria(ruta_id, fecha_reporte):
    """
    Calcula todas las métricas del día para una ruta específica.
    Incluye: abonos (PIX/efectivo), préstamos, caja, gastos, créditos, clientes sin pago.
    """
    from datetime import datetime, timedelta

    # Obtener la ruta y su cobrador
    ruta = Ruta.query.get(ruta_id)
    if not ruta:
        return None

    cobrador = ruta.cobrador

    # Definir el rango de fecha (todo el día)
    fecha_inicio = datetime.combine(fecha_reporte, datetime.min.time())
    fecha_fin = datetime.combine(fecha_reporte, datetime.max.time())

    # ==================== ABONOS DEL DÍA (Pagos de clientes) ====================
    pagos_del_dia = db.session.query(Pago).join(Prestamo).filter(
        Prestamo.ruta_id == ruta_id,
        Pago.fecha_pago >= fecha_inicio,
        Pago.fecha_pago <= fecha_fin
    ).all()

    # Desglose por método de pago
    total_pix = sum(float(p.monto) for p in pagos_del_dia if p.metodo_pago == 'PIX')
    total_efectivo = sum(float(p.monto) for p in pagos_del_dia if p.metodo_pago in ['EFECTIVO', None])
    total_transferencia = sum(float(p.monto) for p in pagos_del_dia if p.metodo_pago == 'TRANSFERENCIA')
    total_abonos = sum(float(p.monto) for p in pagos_del_dia)

    # Detalle de pagos
    detalle_pagos = []
    for p in pagos_del_dia:
        detalle_pagos.append({
            'cliente': p.prestamo.cliente.nombre if p.prestamo and p.prestamo.cliente else 'N/A',
            'monto': float(p.monto),
            'metodo': p.metodo_pago or 'EFECTIVO',
            'hora': p.fecha_pago.strftime('%H:%M') if p.fecha_pago else ''
        })

    # ==================== PRÉSTAMOS DEL DÍA (Desembolsos) ====================
    prestamos_del_dia = Prestamo.query.filter(
        Prestamo.ruta_id == ruta_id,
        Prestamo.fecha_inicio >= fecha_inicio,
        Prestamo.fecha_inicio <= fecha_fin
    ).all()

    total_desembolsos = sum(float(p.monto_prestado) for p in prestamos_del_dia)

    # Detalle de créditos otorgados
    creditos_otorgados = []
    for p in prestamos_del_dia:
        creditos_otorgados.append({
            'cliente': p.cliente.nombre if p.cliente else 'N/A',
            'celular': p.cliente.telefono if p.cliente else '',
            'monto': float(p.monto_prestado),
            'cuotas': p.numero_cuotas,
            'frecuencia': p.frecuencia
        })

    # ==================== GASTOS Y MOVIMIENTOS DEL DÍA ====================
    # Gastos operativos
    gastos_query = Transaccion.query.filter(
        Transaccion.naturaleza == 'EGRESO',
        Transaccion.fecha >= fecha_inicio,
        Transaccion.fecha <= fecha_fin
    )

    # Si hay cobrador asignado, filtrar por él
    if cobrador:
        gastos_query = gastos_query.filter(Transaccion.usuario_origen_id == cobrador.id)

    gastos_del_dia = gastos_query.all()
    total_gastos = sum(float(g.monto) for g in gastos_del_dia)

    detalle_gastos = []
    for g in gastos_del_dia:
        detalle_gastos.append({
            'concepto': g.concepto,
            'descripcion': g.descripcion or '',
            'monto': float(g.monto)
        })

    # Movimientos (traslados)
    movimientos_query = Transaccion.query.filter(
        Transaccion.naturaleza == 'TRASLADO',
        Transaccion.fecha >= fecha_inicio,
        Transaccion.fecha <= fecha_fin
    )

    if cobrador:
        movimientos_query = movimientos_query.filter(
            or_(
                Transaccion.usuario_origen_id == cobrador.id,
                Transaccion.usuario_destino_id == cobrador.id
            )
        )

    movimientos_del_dia = movimientos_query.all()

    detalle_movimientos = []
    for m in movimientos_del_dia:
        es_entrada = m.usuario_destino_id == cobrador.id if cobrador else False
        detalle_movimientos.append({
            'concepto': m.concepto,
            'descripcion': m.descripcion or '',
            'monto': float(m.monto),
            'tipo': 'ENTRADA' if es_entrada else 'SALIDA'
        })

    total_movimientos_entrada = sum(float(m.monto) for m in movimientos_del_dia
                                    if cobrador and m.usuario_destino_id == cobrador.id)
    total_movimientos_salida = sum(float(m.monto) for m in movimientos_del_dia
                                   if cobrador and m.usuario_origen_id == cobrador.id)

    # ==================== CAJA DEL DÍA ====================
    # Caja = Abonos - Desembolsos - Gastos + Traslados Entrada - Traslados Salida
    caja_dia = total_abonos - total_desembolsos - total_gastos + total_movimientos_entrada - total_movimientos_salida

    # ==================== CLIENTES QUE NO PAGARON ====================
    # Obtener todos los préstamos activos de la ruta
    prestamos_activos = Prestamo.query.filter(
        Prestamo.ruta_id == ruta_id,
        Prestamo.estado == 'ACTIVO'
    ).all()

    # IDs de clientes que pagaron hoy
    clientes_que_pagaron_ids = set()
    for p in pagos_del_dia:
        if p.prestamo and p.prestamo.cliente:
            clientes_que_pagaron_ids.add(p.prestamo.cliente_id)

    # Filtrar clientes sin pago que deberían haber pagado (según frecuencia)
    dia_semana = fecha_reporte.weekday()  # 0=Lunes, 6=Domingo

    clientes_sin_pago = []
    for prestamo in prestamos_activos:
        # Verificar si debía pagar ese día según la frecuencia
        debio_pagar = False

        if prestamo.frecuencia == 'DIARIO' and dia_semana != 6:  # No domingo
            debio_pagar = True
        elif prestamo.frecuencia == 'DIARIO_LUNES_VIERNES' and dia_semana < 5:
            debio_pagar = True
        elif prestamo.frecuencia == 'BISEMANAL' and dia_semana in [1, 4]:  # Martes y Viernes
            debio_pagar = True
        elif prestamo.frecuencia == 'SEMANAL':
            # Asumimos que el día de pago es el mismo día de la semana que inició
            if prestamo.fecha_inicio:
                dia_inicio = prestamo.fecha_inicio.weekday() if hasattr(prestamo.fecha_inicio, 'weekday') else 0
                if dia_semana == dia_inicio:
                    debio_pagar = True

        # Si debía pagar pero no pagó
        if debio_pagar and prestamo.cliente_id not in clientes_que_pagaron_ids:
            cliente = prestamo.cliente
            if cliente:
                # Calcular último pago y días de atraso
                ultimo_pago_fecha = None
                dias_atraso = 0
                if prestamo.fecha_ultimo_pago:
                    if hasattr(prestamo.fecha_ultimo_pago, 'date'):
                        ultimo_pago_fecha = prestamo.fecha_ultimo_pago.date()
                    else:
                        ultimo_pago_fecha = prestamo.fecha_ultimo_pago
                    dias_atraso = (fecha_reporte - ultimo_pago_fecha).days
                else:
                    # Si nunca ha pagado, contar desde inicio del préstamo
                    if prestamo.fecha_inicio:
                        fi = prestamo.fecha_inicio.date() if hasattr(prestamo.fecha_inicio, 'date') else prestamo.fecha_inicio
                        dias_atraso = (fecha_reporte - fi).days

                clientes_sin_pago.append({
                    'id': cliente.id,
                    'nombre': cliente.nombre,
                    'telefono': cliente.telefono or '',
                    'saldo_pendiente': float(prestamo.saldo_actual),
                    'cuotas_atrasadas': prestamo.cuotas_atrasadas,
                    'valor_cuota': float(prestamo.valor_cuota) if prestamo.valor_cuota else 0,
                    'ultimo_pago': ultimo_pago_fecha.strftime('%d/%m') if ultimo_pago_fecha else 'Nunca',
                    'dias_atraso': dias_atraso
                })

    # ==================== ESTADÍSTICAS RESUMEN ====================
    total_clientes_activos = len(prestamos_activos)
    clientes_que_pagaron = len(clientes_que_pagaron_ids)
    tasa_cobro = (clientes_que_pagaron / total_clientes_activos * 100) if total_clientes_activos > 0 else 0

    # Por cobrar esperado del día
    por_cobrar_esperado = 0
    for p in prestamos_activos:
        cobra_hoy = False
        if p.frecuencia == 'DIARIO' and dia_semana != 6:
            cobra_hoy = True
        elif p.frecuencia == 'DIARIO_LUNES_VIERNES' and dia_semana < 5:
            cobra_hoy = True
        elif p.frecuencia == 'BISEMANAL' and dia_semana in [1, 4]:
            cobra_hoy = True
        if cobra_hoy:
            por_cobrar_esperado += float(p.valor_cuota) if p.valor_cuota else 0

    tasa_cobro_monto = (total_abonos / por_cobrar_esperado * 100) if por_cobrar_esperado > 0 else 0

    return {
        'ruta': {
            'id': ruta.id,
            'nombre': ruta.nombre,
            'moneda': ruta.moneda or 'COP',
            'simbolo': ruta.simbolo_moneda or '$'
        },
        'cobrador': {
            'id': cobrador.id if cobrador else None,
            'nombre': cobrador.nombre if cobrador else 'Sin asignar'
        },
        'fecha': fecha_reporte.strftime('%d/%m/%Y'),
        'fecha_iso': fecha_reporte.strftime('%Y-%m-%d'),

        # Abonos
        'total_abonos': total_abonos,
        'abonos_pix': total_pix,
        'abonos_efectivo': total_efectivo,
        'abonos_transferencia': total_transferencia,
        'abonos_pix_transferencia': total_pix + total_transferencia,
        'detalle_pagos': detalle_pagos,
        'num_pagos': len(pagos_del_dia),

        # Desembolsos
        'total_desembolsos': total_desembolsos,
        'creditos_otorgados': creditos_otorgados,
        'num_creditos': len(prestamos_del_dia),

        # Gastos y movimientos
        'total_gastos': total_gastos,
        'detalle_gastos': detalle_gastos,
        'detalle_movimientos': detalle_movimientos,
        'total_movimientos_entrada': total_movimientos_entrada,
        'total_movimientos_salida': total_movimientos_salida,

        # Caja
        'caja_dia': caja_dia,

        # Clientes sin pago
        'clientes_sin_pago': clientes_sin_pago,
        'num_clientes_sin_pago': len(clientes_sin_pago),

        # Estadísticas
        'total_clientes_activos': total_clientes_activos,
        'clientes_que_pagaron': clientes_que_pagaron,
        'tasa_cobro_clientes': round(tasa_cobro, 1),
        'por_cobrar_esperado': por_cobrar_esperado,
        'tasa_cobro_monto': round(tasa_cobro_monto, 1)
    }


@reportes_bp.route('/reportes/ruta')
@login_required
def reporte_ruta():
    """Reporte diario por Ruta/Recaudador"""
    usuario_id = session.get('usuario_id')
    rol = session.get('rol')

    # Obtener parámetros
    ruta_id = request.args.get('ruta_id', type=int)
    oficina_id = request.args.get('oficina_id', type=int)
    fecha_str = request.args.get('fecha')

    # Fecha por defecto: ayer
    if fecha_str:
        fecha_reporte = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    else:
        fecha_reporte = (datetime.now() - timedelta(days=1)).date()

    # Cargar rutas y oficinas para filtros
    todas_las_rutas = Ruta.query.filter_by(activo=True).order_by(Ruta.nombre).all()
    todas_las_oficinas = Oficina.query.filter_by(activo=True).order_by(Oficina.nombre).all()

    # Si es cobrador, solo mostrar su ruta
    if rol == 'cobrador':
        todas_las_rutas = [r for r in todas_las_rutas if r.cobrador_id == usuario_id]
        if todas_las_rutas and not ruta_id:
            ruta_id = todas_las_rutas[0].id

    # Si hay oficina seleccionada, filtrar rutas
    if oficina_id:
        todas_las_rutas = [r for r in todas_las_rutas if r.oficina_id == oficina_id]

    # Calcular métricas si hay ruta seleccionada
    metricas = None
    if ruta_id:
        metricas = calcular_metricas_ruta_diaria(ruta_id, fecha_reporte)

    return render_template('reporte_ruta.html',
        metricas=metricas,
        fecha_reporte=fecha_reporte.strftime('%Y-%m-%d'),
        fecha_display=fecha_reporte.strftime('%d/%m/%Y'),
        ruta_seleccionada_id=ruta_id,
        oficina_seleccionada_id=oficina_id,
        todas_las_rutas=todas_las_rutas,
        todas_las_oficinas=todas_las_oficinas,
        nombre=session.get('nombre'),
        rol=rol
    )


def generar_pdf_ruta_reportlab(metricas, fecha_generacion):
    """Genera PDF Dark Mode profesional con ReportLab - Alto contraste y alineación perfecta."""
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib.colors import HexColor
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

    # ============================================================
    # PALETA DE COLORES DARK MODE
    # ============================================================
    BG_PAGE = HexColor('#1E1E1E')       # Fondo de página
    BG_SECTION = HexColor('#2C2C2C')    # Fondo de secciones
    BG_HEADER = HexColor('#252525')     # Fondo header tabla
    BG_ROW_ALT = HexColor('#333333')    # Fila alternada
    CYAN = HexColor('#00BCD4')          # Acento principal
    GREEN = HexColor('#00E676')         # Positivo / Efectivo
    RED = HexColor('#FF5252')           # Negativo / Sin pago
    GOLD = HexColor('#FFD740')          # Destacado
    WHITE = HexColor('#FFFFFF')         # Texto principal
    LIGHT_GRAY = HexColor('#DDDDDD')    # Texto etiquetas
    MID_GRAY = HexColor('#AAAAAA')      # Texto secundario
    BORDER = HexColor('#555555')        # Líneas de tabla
    BORDER_SUBTLE = HexColor('#444444') # Líneas sutiles

    simbolo = metricas['ruta']['simbolo']
    def fmt(n):
        return f"{simbolo}{n:,.0f}"

    # ============================================================
    # ESTILOS DE PÁRRAFO (Definidos al inicio para consistencia)
    # ============================================================
    # --- Header ---
    sTitle = ParagraphStyle('sTitle', fontName='Helvetica-Bold', fontSize=20, textColor=CYAN, alignment=TA_LEFT, leading=24)
    sSub = ParagraphStyle('sSub', fontName='Helvetica', fontSize=9, textColor=MID_GRAY, alignment=TA_LEFT)
    sFecha = ParagraphStyle('sFecha', fontName='Helvetica-Bold', fontSize=14, textColor=CYAN, alignment=TA_RIGHT, leading=18)
    sFechaLbl = ParagraphStyle('sFechaLbl', fontName='Helvetica', fontSize=8, textColor=MID_GRAY, alignment=TA_RIGHT)

    # --- Info Ruta ---
    sRuta = ParagraphStyle('sRuta', fontName='Helvetica-Bold', fontSize=15, textColor=GREEN, alignment=TA_LEFT, leading=18)
    sCobrador = ParagraphStyle('sCobrador', fontName='Helvetica', fontSize=10, textColor=LIGHT_GRAY, alignment=TA_LEFT)

    # --- Títulos de sección ---
    sSecTitle = ParagraphStyle('sSecTitle', fontName='Helvetica-Bold', fontSize=11, textColor=CYAN, alignment=TA_LEFT, leading=14)

    # --- Cuadre de Ruta ---
    sCuadreTitle = ParagraphStyle('sCuadreTitle', fontName='Helvetica-Bold', fontSize=13, textColor=CYAN, alignment=TA_CENTER, leading=16)
    sLabelL = ParagraphStyle('sLabelL', fontName='Helvetica', fontSize=11, textColor=WHITE, alignment=TA_LEFT, leading=14)
    sValGreen = ParagraphStyle('sValGreen', fontName='Helvetica-Bold', fontSize=12, textColor=GREEN, alignment=TA_RIGHT, leading=15)
    sValRed = ParagraphStyle('sValRed', fontName='Helvetica-Bold', fontSize=12, textColor=RED, alignment=TA_RIGHT, leading=15)
    sValWhite = ParagraphStyle('sValWhite', fontName='Helvetica-Bold', fontSize=12, textColor=WHITE, alignment=TA_RIGHT, leading=15)
    sTotalLbl = ParagraphStyle('sTotalLbl', fontName='Helvetica-Bold', fontSize=11, textColor=CYAN, alignment=TA_CENTER)
    sTotalPos = ParagraphStyle('sTotalPos', fontName='Helvetica-Bold', fontSize=24, textColor=GREEN, alignment=TA_CENTER, leading=28)
    sTotalNeg = ParagraphStyle('sTotalNeg', fontName='Helvetica-Bold', fontSize=24, textColor=RED, alignment=TA_CENTER, leading=28)

    # --- Stats boxes ---
    sStatGreen = ParagraphStyle('sStatGreen', fontName='Helvetica-Bold', fontSize=18, textColor=GREEN, alignment=TA_CENTER, leading=22)
    sStatCyan = ParagraphStyle('sStatCyan', fontName='Helvetica-Bold', fontSize=18, textColor=CYAN, alignment=TA_CENTER, leading=22)
    sStatLbl = ParagraphStyle('sStatLbl', fontName='Helvetica', fontSize=9, textColor=LIGHT_GRAY, alignment=TA_CENTER)

    # --- Tablas de datos ---
    sTh = ParagraphStyle('sTh', fontName='Helvetica-Bold', fontSize=9, textColor=CYAN, alignment=TA_LEFT, leading=12)
    sThR = ParagraphStyle('sThR', fontName='Helvetica-Bold', fontSize=9, textColor=CYAN, alignment=TA_RIGHT, leading=12)
    sTd = ParagraphStyle('sTd', fontName='Helvetica', fontSize=10, textColor=WHITE, alignment=TA_LEFT, leading=13)
    sTdBold = ParagraphStyle('sTdBold', fontName='Helvetica-Bold', fontSize=10, textColor=WHITE, alignment=TA_LEFT, leading=13)
    sTdR = ParagraphStyle('sTdR', fontName='Helvetica', fontSize=10, textColor=WHITE, alignment=TA_RIGHT, leading=13)
    sTdRed = ParagraphStyle('sTdRed', fontName='Helvetica-Bold', fontSize=10, textColor=RED, alignment=TA_RIGHT, leading=13)
    sTdGreen = ParagraphStyle('sTdGreen', fontName='Helvetica-Bold', fontSize=10, textColor=GREEN, alignment=TA_RIGHT, leading=13)
    sTdGray = ParagraphStyle('sTdGray', fontName='Helvetica', fontSize=10, textColor=MID_GRAY, alignment=TA_LEFT, leading=13)

    # --- Footer / No data ---
    sFooter = ParagraphStyle('sFooter', fontName='Helvetica', fontSize=8, textColor=MID_GRAY, alignment=TA_CENTER)
    sNoData = ParagraphStyle('sNoData', fontName='Helvetica-Oblique', fontSize=10, textColor=MID_GRAY, alignment=TA_CENTER)
    sNoDataOk = ParagraphStyle('sNoDataOk', fontName='Helvetica-Bold', fontSize=10, textColor=GREEN, alignment=TA_CENTER)

    # ============================================================
    # HELPERS para estilos de tabla reutilizables
    # ============================================================
    def data_table_style(num_rows, has_data=True):
        """Retorna TableStyle para tablas de datos con header oscuro y filas alternadas."""
        style = [
            ('BACKGROUND', (0, 0), (-1, 0), BG_HEADER),
            ('TEXTCOLOR', (0, 0), (-1, 0), CYAN),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, CYAN),
            ('TOPPADDING', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        if has_data:
            for i in range(1, num_rows):
                bg = BG_SECTION if i % 2 == 1 else BG_ROW_ALT
                style.append(('BACKGROUND', (0, i), (-1, i), bg))
                if i < num_rows - 1:
                    style.append(('LINEBELOW', (0, i), (-1, i), 0.5, BORDER_SUBTLE))
        else:
            style.append(('BACKGROUND', (0, 1), (-1, -1), BG_SECTION))
        return TableStyle(style)

    def section_wrapper_style():
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), BG_SECTION),
            ('BOX', (0, 0), (-1, -1), 0.5, BORDER),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ('LINEBELOW', (0, 0), (-1, 0), 1, CYAN),
        ])

    # ============================================================
    # CONSTRUIR DOCUMENTO
    # ============================================================
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=1.2*cm, rightMargin=1.2*cm,
                            topMargin=1*cm, bottomMargin=1*cm)
    pw = A4[0] - 2.4 * cm  # ancho útil
    elements = []

    def draw_bg(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(BG_PAGE)
        canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)
        canvas.restoreState()

    # ============================================================
    # 1. HEADER
    # ============================================================
    hdr = Table([[
        [Paragraph("&#9670; DIAMANTE PRO", sTitle),
         Paragraph("Sistema de Gesti&oacute;n de Cr&eacute;ditos", sSub)],
        [Paragraph(metricas['fecha'], sFecha),
         Paragraph("Fecha del Reporte", sFechaLbl)]
    ]], colWidths=[pw * 0.6, pw * 0.4])
    hdr.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BG_SECTION),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (0, 0), 14),
        ('RIGHTPADDING', (1, 0), (1, 0), 14),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LINEBEFORE', (0, 0), (0, -1), 4, CYAN),
    ]))
    elements.append(hdr)
    elements.append(Spacer(1, 10))

    # ============================================================
    # 2. RUTA + COBRADOR
    # ============================================================
    elements.append(Paragraph(metricas['ruta']['nombre'], sRuta))
    elements.append(Paragraph(
        f"Cobrador: <b><font color='#FFFFFF'>{metricas['cobrador']['nombre']}</font></b>", sCobrador))
    elements.append(Spacer(1, 12))

    # ============================================================
    # 3. CUADRE DE RUTA
    # ============================================================
    cuadre_data = [
        [Paragraph("Total Abonos (Cobros)", sLabelL),
         Paragraph(fmt(metricas['total_abonos']), sValGreen)],
        [Paragraph("(-) Desembolsos (Pr\u00e9stamos)", sLabelL),
         Paragraph(fmt(metricas['total_desembolsos']), sValRed)],
        [Paragraph("(-) Gastos", sLabelL),
         Paragraph(fmt(metricas['total_gastos']), sValRed)],
    ]
    if metricas['total_movimientos_entrada'] > 0:
        cuadre_data.append([Paragraph("(+) Traslados Recibidos", sLabelL),
                            Paragraph(fmt(metricas['total_movimientos_entrada']), sValGreen)])
    if metricas['total_movimientos_salida'] > 0:
        cuadre_data.append([Paragraph("(-) Traslados Enviados", sLabelL),
                            Paragraph(fmt(metricas['total_movimientos_salida']), sValRed)])

    iw = pw * 0.75
    cuadre_tbl = Table(cuadre_data, colWidths=[iw * 0.62, iw * 0.38])
    cuadre_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BG_HEADER),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (0, -1), 14),
        ('RIGHTPADDING', (1, 0), (1, -1), 14),
        ('LINEBELOW', (0, 0), (-1, -2), 0.5, BORDER_SUBTLE),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    total_s = sTotalPos if metricas['caja_dia'] >= 0 else sTotalNeg
    total_tbl = Table([
        [Paragraph("TOTAL CAJA", sTotalLbl)],
        [Paragraph(fmt(metricas['caja_dia']), total_s)],
    ], colWidths=[iw])
    total_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BG_PAGE),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LINEABOVE', (0, 0), (-1, 0), 2, CYAN),
    ]))

    cuadre_wrap = Table([
        [Paragraph("CUADRE DE RUTA", sCuadreTitle)],
        [cuadre_tbl],
        [total_tbl],
    ], colWidths=[pw])
    cuadre_wrap.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BG_HEADER),
        ('BOX', (0, 0), (-1, -1), 2, CYAN),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    elements.append(cuadre_wrap)
    elements.append(Spacer(1, 12))

    # ============================================================
    # 4. ABONOS DEL DIA
    # ============================================================
    stat_row = [[
        [Paragraph(fmt(metricas['abonos_efectivo']), sStatGreen),
         Paragraph("Efectivo", sStatLbl)],
        [Paragraph(fmt(metricas.get('abonos_pix_transferencia', 0)), sStatCyan),
         Paragraph("PIX / Transferencia", sStatLbl)],
    ]]
    stat_tbl = Table(stat_row, colWidths=[pw * 0.46, pw * 0.46], hAlign='CENTER')
    stat_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), HexColor('#1B3D2A')),
        ('BACKGROUND', (1, 0), (1, 0), HexColor('#1A2D3D')),
        ('BOX', (0, 0), (0, 0), 2, GREEN),
        ('BOX', (1, 0), (1, 0), 2, CYAN),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    abonos_wrap = Table([
        [Paragraph(f"ABONOS DEL D\u00cdA ({metricas['num_pagos']} pagos)", sSecTitle)],
        [Spacer(1, 4)],
        [stat_tbl],
    ], colWidths=[pw])
    abonos_wrap.setStyle(section_wrapper_style())
    elements.append(abonos_wrap)
    elements.append(Spacer(1, 10))

    # ============================================================
    # 5. CREDITOS OTORGADOS
    # ============================================================
    cr_rows = [[Paragraph("Cliente", sTh), Paragraph("Celular", sTh), Paragraph("Valor", sThR)]]
    has_cred = bool(metricas['creditos_otorgados'])
    if has_cred:
        for c in metricas['creditos_otorgados']:
            cr_rows.append([Paragraph(c['cliente'], sTd), Paragraph(c['celular'], sTdGray), Paragraph(fmt(c['monto']), sTdR)])
    else:
        cr_rows.append([Paragraph("Sin cr\u00e9ditos otorgados", sNoData), '', ''])

    cr_tbl = Table(cr_rows, colWidths=[pw * 0.40, pw * 0.30, pw * 0.25])
    cr_style = data_table_style(len(cr_rows), has_cred)
    if not has_cred:
        cr_style.add('SPAN', (0, 1), (-1, 1))
    cr_tbl.setStyle(cr_style)

    stat_des = Table([
        [Paragraph(fmt(metricas['total_desembolsos']), sStatCyan)],
        [Paragraph("Total Desembolsado", sStatLbl)],
    ], colWidths=[pw * 0.6], hAlign='CENTER')
    stat_des.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), HexColor('#1A2D3D')),
        ('BOX', (0, 0), (-1, -1), 1, CYAN),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))

    cred_wrap = Table([
        [Paragraph(f"CR\u00c9DITOS OTORGADOS ({metricas['num_creditos']})", sSecTitle)],
        [Spacer(1, 4)],
        [stat_des],
        [Spacer(1, 4)],
        [cr_tbl],
    ], colWidths=[pw])
    cred_wrap.setStyle(section_wrapper_style())
    elements.append(cred_wrap)
    elements.append(Spacer(1, 10))

    # ============================================================
    # 6. GASTOS Y MOVIMIENTOS
    # ============================================================
    g_rows = [[Paragraph("Concepto", sTh), Paragraph("Descripci\u00f3n", sTh), Paragraph("Monto", sThR)]]
    has_g = bool(metricas['detalle_gastos'] or metricas['detalle_movimientos'])
    if has_g:
        for g in metricas['detalle_gastos']:
            g_rows.append([Paragraph(g['concepto'], sTd), Paragraph(g['descripcion'], sTdGray),
                           Paragraph(f"-{fmt(g['monto'])}", sTdRed)])
        for m in metricas['detalle_movimientos']:
            cs = sTdGreen if m['tipo'] == 'ENTRADA' else sTdRed
            pf = "+" if m['tipo'] == 'ENTRADA' else "-"
            g_rows.append([Paragraph(m['concepto'], sTd), Paragraph(m['descripcion'], sTdGray),
                           Paragraph(f"{pf}{fmt(m['monto'])}", cs)])
    else:
        g_rows.append([Paragraph("Sin gastos ni movimientos", sNoData), '', ''])

    g_tbl = Table(g_rows, colWidths=[pw * 0.28, pw * 0.47, pw * 0.20])
    g_style = data_table_style(len(g_rows), has_g)
    if not has_g:
        g_style.add('SPAN', (0, 1), (-1, 1))
    g_tbl.setStyle(g_style)

    gastos_wrap = Table([
        [Paragraph("GASTOS Y MOVIMIENTOS", sSecTitle)],
        [Spacer(1, 4)],
        [g_tbl],
    ], colWidths=[pw])
    gastos_wrap.setStyle(section_wrapper_style())
    elements.append(gastos_wrap)
    elements.append(Spacer(1, 10))

    # ============================================================
    # 7. CLIENTES SIN PAGO (5 columnas: #, Cliente, Teléfono, Último Pago, Días Atraso)
    # ============================================================
    # Estilos centrados para las columnas de Teléfono, Último Pago y Días Atraso
    sThC = ParagraphStyle('sThC', fontName='Helvetica-Bold', fontSize=9, textColor=CYAN, alignment=TA_CENTER, leading=12)
    sTdC = ParagraphStyle('sTdC', fontName='Helvetica', fontSize=10, textColor=WHITE, alignment=TA_CENTER, leading=13)
    sTdCGray = ParagraphStyle('sTdCGray', fontName='Helvetica', fontSize=10, textColor=MID_GRAY, alignment=TA_CENTER, leading=13)
    sTdCRed = ParagraphStyle('sTdCRed', fontName='Helvetica-Bold', fontSize=10, textColor=RED, alignment=TA_CENTER, leading=13)

    sp_rows = [[
        Paragraph("#", sThC),
        Paragraph("Cliente", sTh),
        Paragraph("Tel\u00e9fono", sThC),
        Paragraph("\u00daltimo Pago", sThC),
        Paragraph("D\u00edas Atraso", sThC),
    ]]
    has_sp = bool(metricas['clientes_sin_pago'])
    if has_sp:
        for i, c in enumerate(metricas['clientes_sin_pago'], 1):
            dias = c.get('dias_atraso', 0)
            # Color de días: rojo si >= 3, amarillo si 1-2, gris si 0
            if dias >= 3:
                dias_style = sTdCRed
            elif dias >= 1:
                dias_p_style = ParagraphStyle('sDiasWarn', fontName='Helvetica-Bold', fontSize=10, textColor=GOLD, alignment=TA_CENTER, leading=13)
                dias_style = dias_p_style
            else:
                dias_style = sTdC

            sp_rows.append([
                Paragraph(str(i), sTdC),
                Paragraph(c['nombre'], sTdBold),
                Paragraph(c['telefono'], sTdCGray),
                Paragraph(c.get('ultimo_pago', '—'), sTdC),
                Paragraph(str(dias), dias_style),
            ])
    else:
        sp_rows.append([Paragraph("Todos los clientes pagaron", sNoDataOk), '', '', '', ''])

    sp_tbl = Table(sp_rows, colWidths=[pw * 0.06, pw * 0.34, pw * 0.24, pw * 0.18, pw * 0.18])
    sp_style = data_table_style(len(sp_rows), has_sp)
    if not has_sp:
        sp_style.add('SPAN', (0, 1), (-1, 1))
    sp_tbl.setStyle(sp_style)

    sp_wrap = Table([
        [Paragraph(f"CLIENTES SIN PAGO ({metricas['num_clientes_sin_pago']})", sSecTitle)],
        [Spacer(1, 4)],
        [sp_tbl],
    ], colWidths=[pw])
    sp_wrap.setStyle(section_wrapper_style())
    elements.append(sp_wrap)
    elements.append(Spacer(1, 15))

    # ============================================================
    # 8. FOOTER
    # ============================================================
    elements.append(Paragraph(
        f"Reporte generado el {fecha_generacion}  |  DIAMANTE PRO  |  Sistema de Gesti\u00f3n de Cr\u00e9ditos",
        sFooter))

    # ============================================================
    # BUILD
    # ============================================================
    doc.build(elements, onFirstPage=draw_bg, onLaterPages=draw_bg)
    return buffer.getvalue()


@reportes_bp.route('/reportes/ruta/pdf')
@login_required
def reporte_ruta_pdf():
    """Genera PDF del reporte por ruta con ReportLab."""
    ruta_id = request.args.get('ruta_id', type=int)
    fecha_str = request.args.get('fecha')

    if not ruta_id:
        return jsonify({'error': 'Debe seleccionar una ruta'}), 400

    if fecha_str:
        fecha_reporte = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    else:
        fecha_reporte = (datetime.now() - timedelta(days=1)).date()

    metricas = calcular_metricas_ruta_diaria(ruta_id, fecha_reporte)
    if not metricas:
        return jsonify({'error': 'Ruta no encontrada'}), 404

    fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M')
    pdf = generar_pdf_ruta_reportlab(metricas, fecha_generacion)

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    nombre_ruta = metricas['ruta']['nombre'].replace(' ', '_')
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_{nombre_ruta}_{fecha_reporte.strftime("%Y%m%d")}.pdf'
    return response


@reportes_bp.route('/api/reportes/ruta/metricas')
@login_required
def api_metricas_ruta():
    """API para obtener métricas de ruta en tiempo real"""
    ruta_id = request.args.get('ruta_id', type=int)
    fecha_str = request.args.get('fecha')

    if not ruta_id:
        return jsonify({'error': 'Debe especificar ruta_id'}), 400

    if fecha_str:
        fecha_reporte = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    else:
        fecha_reporte = (datetime.now() - timedelta(days=1)).date()

    metricas = calcular_metricas_ruta_diaria(ruta_id, fecha_reporte)

    if not metricas:
        return jsonify({'error': 'Ruta no encontrada'}), 404

    return jsonify(metricas)
